import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, Border

# ファイルアップロード
uploaded_file = st.file_uploader("Excelファイルをアップロードしてください", type=["xlsx"])

if uploaded_file is not None:
    try:
        df = pd.read_excel(uploaded_file, engine='openpyxl')
        st.success("ファイルを読み込みました")
        st.write(df.head())  # プレビュー表示

        # 一時的にファイルを保存
        input_file_path = "uploaded_input.xlsx"
        with open(input_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        output_file_path = "output.xlsx"

        # Excelファイルを処理
        xls = pd.ExcelFile(input_file_path, engine="openpyxl")
        df_classification = pd.read_excel(xls, sheet_name="クラス分け")
        df_mixture = pd.read_excel(xls, sheet_name="配合表")

        df_mixture.rename(columns={'Unnamed: 0': 'TrendFactor'}, inplace=True)
        df_classified_mixture = df_classification.merge(df_mixture, on="TrendFactor", how="left")

        time_order = ["本来性", "現状", "志向性"]
        df_classified_mixture["time_category"] = df_classified_mixture["TrendFactor"].str.extract(f"({'|'.join(time_order)})")[0]
        df_classified_mixture["time_category"] = pd.Categorical(df_classified_mixture["time_category"], categories=time_order, ordered=True)
        df_classified_mixture = df_classified_mixture.sort_values(by=["class", "time_category", "TrendFactor"])

        df_classified_mixture.set_index(["class", "TrendFactor"], inplace=True)
        df_classified_mixture = df_classified_mixture.reset_index()

        df_class_time_sum = df_classified_mixture.groupby(["class", "time_category"]).sum(numeric_only=True).reset_index()
        df_class_time_sum["time_category"] = pd.Categorical(df_class_time_sum["time_category"], categories=time_order, ordered=True)
        df_class_time_sum = df_class_time_sum.sort_values(by=["class", "time_category"])

        numeric_columns = df_class_time_sum.select_dtypes(include=[np.number]).columns
        df_class_time_sum["BF合計"] = np.sqrt(df_class_time_sum[numeric_columns].sum(axis=1))

        df_class_time_normalized = df_class_time_sum.copy()
        df_class_time_normalized.drop(columns=["BF合計"], inplace=True)
        df_class_time_normalized[numeric_columns] = df_class_time_sum[numeric_columns].div(df_class_time_sum["BF合計"], axis=0)
        df_class_time_normalized = df_class_time_normalized.sort_values(by=["class", "time_category"])

        df_class_time_normalized[numeric_columns] = df_class_time_normalized[numeric_columns].applymap(lambda x: 0 if x < 0.1 else x)

        def is_equal(a, b):
            return abs(a - b) < 0.1 or (max(a, b) / min(a, b) < 1.5 if min(a, b) > 0 else False)

        def determine_bf_type(bf_h, bf_c, bf_s):
            if bf_h == 0 and bf_c == 0 and bf_s == 0:
                return "none"
            if bf_h != 0 and bf_c == 0 and bf_s == 0:
                return "d0-1"
            if bf_s != 0 and bf_h == 0 and bf_c == 0:
                return "d0-2"
            if bf_c != 0 and bf_h == 0 and bf_s == 0:
                return "d0-3"
            if bf_c == 0 and is_equal(bf_h, bf_s):
                return "d1-eq"
            if bf_c == 0 and bf_h > bf_s:
                return "d1->"
            if bf_c == 0 and bf_h < bf_s:
                return "d1-<"
            if bf_h == 0 and bf_c != 0 and bf_s != 0:
                return "d2"
            if bf_s == 0 and bf_c != 0 and bf_h != 0:
                return "d3"
            if is_equal(bf_h, bf_s) and (bf_c == bf_h or bf_c == bf_s or bf_c > bf_h or bf_c > bf_s):
                return "s1"
            if bf_h < bf_s:
                return "d2"
            if bf_h > bf_s:
                return "d3"
            return "unknown"

        df_bf_type = df_class_time_normalized.pivot(index="class", columns="time_category", values=numeric_columns)
        df_bf_type.columns = [f"{col[1]}_{col[0]}" for col in df_bf_type.columns]
        df_bf_type.reset_index(drop=False, inplace=True)
        df_bf_type = df_bf_type.loc[:, ~df_bf_type.columns.duplicated()]

        for col in numeric_columns:
            df_bf_type[col] = df_bf_type.apply(lambda row: determine_bf_type(row.get(f"本来性_{col}", 0),
                                                                             row.get(f"現状_{col}", 0),
                                                                             row.get(f"志向性_{col}", 0)), axis=1)

        bf_type_columns = ["class"] + [col for col in numeric_columns]
        df_bf_type = df_bf_type[bf_type_columns]

        # Excel保存
        with pd.ExcelWriter(output_file_path, engine="openpyxl", mode="w") as writer:
            df_classified_mixture.to_excel(writer, sheet_name="クラス別配合表", index=False)
            df_class_time_sum.to_excel(writer, sheet_name="クラス別配合時制別合計表", index=False)
            df_class_time_normalized.to_excel(writer, sheet_name="クラス別時制別正規化配合表", index=False)
            df_bf_type.to_excel(writer, sheet_name="クラス別BFタイプ", index=False)

        # ヘッダーを標準フォントに
        wb = load_workbook(output_file_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.font = Font(bold=False)
                cell.border = Border()
        wb.save(output_file_path)

        # ダウンロードボタン
        with open(output_file_path, "rb") as f:
            st.download_button("修正済みExcelをダウンロード", f, file_name="output.xlsx")

    except Exception as e:
        st.error(f"ファイルの処理中にエラーが発生しました: {e}")
